"""
lambda_function.py
------------------------------------------------------------------
Backend for the Lake America Family Physicians chat widget.

What this does, end to end, on every request:
  1. Receives {sessionId, messages} from the widget via API Gateway.
  2. Calls the Anthropic API server-side (API key never touches the browser).
  3. Saves the running conversation to DynamoDB (survives Lambda cold starts,
     lets you look up any conversation later, auto-expires via TTL).
  4. Scans the conversation for a captured new-patient lead (phone number
     pattern). The first time one is found in a session, asks Claude to pull
     out structured fields (name, phone, request type, details) and appends
     them as a new row to an Excel workbook stored in S3 — the practice's
     running new-patient intake sheet.
  5. Returns { reply: "<assistant text>" } to the widget.

This intentionally does NOT touch any practice-management/scheduling system —
per the real constraint we identified (Lake America's site vendor, Rimage,
has no public booking API). This is the realistic v1: capture + log to a
spreadsheet, not live calendar booking.

ENVIRONMENT VARIABLES (set these when deploying — see deploy.sh):
  ANTHROPIC_API_KEY   - your Anthropic API key (server-side only)
  DYNAMODB_TABLE      - name of the conversations table
  S3_BUCKET           - bucket holding the intake workbook
  S3_KEY              - object key of the workbook, e.g. new-patient-intake.xlsx
  ALLOWED_ORIGIN      - the origin allowed to call this API (CORS), e.g.
                        https://lakeamerica.com  (use "*" only for testing)
"""

import json
import os
import re
import time
import uuid
import io
import urllib.request
import urllib.error

import boto3
from openpyxl import Workbook, load_workbook

dynamodb = boto3.resource("dynamodb")
s3 = boto3.client("s3")

TABLE_NAME = os.environ.get("DYNAMODB_TABLE", "chatbot-conversations")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
S3_BUCKET = os.environ.get("S3_BUCKET", "")
S3_KEY = os.environ.get("S3_KEY", "new-patient-intake.xlsx")
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")

ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_MODEL = "claude-sonnet-4-6"

# Placeholder doctor schedule — swap for the real roster before using this
# with an actual client. Keeping it as one clearly-marked block makes that
# a one-line-per-doctor edit later.
DOCTOR_SCHEDULE = """- Dr. Maria Santos — available Monday, Wednesday, Friday
- Dr. James Whitfield — available Tuesday, Thursday, Friday
- Dr. Priya Nair — available Monday, Tuesday, Thursday"""

SYSTEM_PROMPT = f"""You are the front-desk chat assistant embedded on the Lake America Family Physicians website, a small single-location family medicine practice in Clermont, FL (301 Frontage Road, Suite G).

Real facts about the practice:
- Hours: Monday-Friday 8am-5pm, closed weekends
- Phone: (352) 432-3939, Fax: (352) 432-3908
- Doctor weekly availability (PLACEHOLDER NAMES — replace with real roster before real use):
{DOCTOR_SCHEDULE}
- They offer both in-clinic and telemedicine visits
- Affiliated with South Lake Hospital
- Accepted insurance includes: AdventHealth, Aetna, BCBS, CarePlus, Cigna, Cigna Disney, Core Source/Orlando Health, Devoted, Emblem, Florida Blue, GHI, Humana, Medicare, Oscar, Tricare, United Health Care
- No online patient portal or self-service scheduling exists — requests currently go through phone or several specific emails
- New patients establish care by calling or emailing the front desk

Your job:
1. Be the single simple front door — ask what they need (new appointment, existing patient appointment, refill, billing/insurance question, or general question) and route accordingly.
2. Answer general questions directly using the facts above: hours, insurance accepted, doctor weekly availability, telemedicine option, location.
3. You do NOT have access to a live calendar. Capture requests, don't confirm bookings — say staff will confirm the specific time.
4. For a new appointment: ask new/existing patient, in-clinic or telemedicine, reason for visit (general terms only, e.g. "annual checkup", "follow-up" — not symptom detail), preferred days/times, then full name and phone number. Once you have name + phone, confirm back to them that their request has been logged and staff will follow up.
5. For refills: ask medication name and pharmacy.
6. For billing/insurance: ask briefly what the issue is, say it will be routed to billing.
7. NEVER give medical advice, diagnoses, or medication guidance.
8. If the person describes urgent/emergency symptoms, immediately tell them to call 911 or go to the nearest ER, and do not continue collecting appointment details in that message.
9. Warm, concise, professional. 2-4 sentences per reply."""

PHONE_RE = re.compile(r"(\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}")

EXTRACTION_PROMPT = """Read the conversation transcript below between a medical practice's chat assistant and a website visitor. Extract the visitor's intake details as strict JSON with exactly these keys:

{"name": "...", "phone": "...", "request_type": "new appointment | existing patient | refill | billing | other", "details": "short one-line summary"}

If a field genuinely isn't present, use an empty string for it. Return ONLY the JSON object, no other text.

TRANSCRIPT:
"""


def _cors_headers():
    return {
        "Access-Control-Allow-Origin": ALLOWED_ORIGIN,
        "Access-Control-Allow-Headers": "Content-Type",
        "Access-Control-Allow-Methods": "OPTIONS,POST",
        "Content-Type": "application/json",
    }


def _response(status, body_dict):
    return {
        "statusCode": status,
        "headers": _cors_headers(),
        "body": json.dumps(body_dict),
    }


def _call_anthropic_raw(system_prompt, messages, max_tokens=300):
    payload = json.dumps(
        {
            "model": ANTHROPIC_MODEL,
            "max_tokens": max_tokens,
            "system": system_prompt,
            "messages": messages,
        }
    ).encode("utf-8")

    req = urllib.request.Request(
        ANTHROPIC_URL,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8")
        raise RuntimeError(f"Anthropic API error {e.code}: {err_body}")

    for block in data.get("content", []):
        if block.get("type") == "text":
            return block["text"]
    return ""


def call_anthropic(messages):
    reply = _call_anthropic_raw(SYSTEM_PROMPT, messages, max_tokens=300)
    return reply or "Sorry, could you try that again?"


def extract_intake_fields(messages):
    """Ask Claude to turn the raw conversation into structured fields we can
    drop straight into a spreadsheet row. Small, cheap, separate call —
    keeps the main conversational system prompt free of JSON-formatting
    instructions that would make replies feel robotic."""
    transcript = "\n".join(f"{m['role']}: {m['content']}" for m in messages)
    raw = _call_anthropic_raw(
        "You extract structured data from conversations. Reply with ONLY valid JSON, nothing else.",
        [{"role": "user", "content": EXTRACTION_PROMPT + transcript}],
        max_tokens=200,
    )
    try:
        # Strip any accidental markdown code fences before parsing
        cleaned = raw.strip().strip("`").replace("json\n", "", 1) if raw.strip().startswith("```") else raw
        return json.loads(cleaned)
    except (json.JSONDecodeError, AttributeError):
        return {"name": "", "phone": "", "request_type": "other", "details": raw[:200]}


def save_conversation(session_id, messages, lead_logged):
    table = dynamodb.Table(TABLE_NAME)
    table.put_item(
        Item={
            "sessionId": session_id,
            "messages": json.dumps(messages),
            "leadLogged": lead_logged,
            "updatedAt": int(time.time()),
            # TTL: auto-delete conversation rows after 30 days
            "ttl": int(time.time()) + 30 * 24 * 60 * 60,
        }
    )


def get_conversation_meta(session_id):
    table = dynamodb.Table(TABLE_NAME)
    resp = table.get_item(Key={"sessionId": session_id})
    return resp.get("Item")


def append_row_to_excel(row_values):
    """Downloads the current intake workbook from S3 (creates one with
    headers if it doesn't exist yet), appends one row, re-uploads it."""
    if not S3_BUCKET:
        return  # Not configured — skip silently rather than error out

    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
        wb = load_workbook(io.BytesIO(obj["Body"].read()))
        ws = wb.active
    except s3.exceptions.NoSuchKey:
        wb = Workbook()
        ws = wb.active
        ws.title = "New Patient Intake"
        ws.append(["Timestamp", "Name", "Phone", "Request Type", "Details", "Session ID"])

    ws.append(row_values)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3.put_object(
        Bucket=S3_BUCKET,
        Key=S3_KEY,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def maybe_log_lead(session_id, messages, already_logged):
    """If a phone number shows up in the conversation and we haven't already
    logged this session to the spreadsheet, extract structured fields and
    append a row."""
    if already_logged:
        return already_logged

    full_text = "\n".join(f"{m['role']}: {m['content']}" for m in messages)
    if not PHONE_RE.search(full_text):
        return already_logged

    fields = extract_intake_fields(messages)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())

    append_row_to_excel(
        [
            timestamp,
            fields.get("name", ""),
            fields.get("phone", ""),
            fields.get("request_type", ""),
            fields.get("details", ""),
            session_id,
        ]
    )
    return True


def lambda_handler(event, context):
    # API Gateway HTTP API sends OPTIONS preflight through if CORS isn't
    # fully configured at the API level — handle it defensively here too.
    method = event.get("requestContext", {}).get("http", {}).get("method", "POST")
    if method == "OPTIONS":
        return _response(200, {})

    try:
        body = json.loads(event.get("body") or "{}")
    except json.JSONDecodeError:
        return _response(400, {"error": "Invalid JSON body"})

    messages = body.get("messages")
    session_id = body.get("sessionId") or str(uuid.uuid4())

    if not messages or not isinstance(messages, list):
        return _response(400, {"error": "messages[] is required"})

    if not ANTHROPIC_API_KEY:
        return _response(500, {"error": "Server misconfigured: missing ANTHROPIC_API_KEY"})

    try:
        reply_text = call_anthropic(messages)
    except Exception as e:
        return _response(502, {"error": f"Upstream AI call failed: {str(e)}"})

    full_messages = messages + [{"role": "assistant", "content": reply_text}]

    existing = get_conversation_meta(session_id) or {}
    already_logged = bool(existing.get("leadLogged", False))

    try:
        logged_now = maybe_log_lead(session_id, full_messages, already_logged)
    except Exception:
        # Never let a spreadsheet-write failure break the chat response itself
        logged_now = already_logged

    try:
        save_conversation(session_id, full_messages, logged_now)
    except Exception:
        # Never let a DynamoDB failure break the chat response itself
        pass

    return _response(200, {"reply": reply_text, "sessionId": session_id})
